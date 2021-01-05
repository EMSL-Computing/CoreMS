from collections import namedtuple
from openpyxl import load_workbook

def get_emsl_jgi_mapping(wb):

    last_sheet = wb.sheetnames[-1]
    emsl_jgi_mapping = wb[last_sheet]
    emsl_proposals = emsl_jgi_mapping['B']
    jgi_proposals = emsl_jgi_mapping['A']
    emsl_jgi_dict = {}
    for x in range(len(emsl_proposals)): 
        emsl_jgi_dict[emsl_proposals[x].value] = jgi_proposals[x].value
    return emsl_jgi_dict

def get_data_mapping(wb, emsl_jgi_dict):

    first_sheet = wb.sheetnames[0]

    full_list_worksheet = wb[first_sheet]

    emsl_proposal = full_list_worksheet['A'] 
    dataset_id = full_list_worksheet['B'] 
    dataset_name = full_list_worksheet['C'] 
    experiment_id = full_list_worksheet['S'] 
    data_dict = {}

    for x in range(len(dataset_id)): 
        
        data = {
                'data_id' : dataset_id[x].value,  
                'experiment_id' : experiment_id[x].value,
                'emsl_proposal_id' : emsl_proposal[x].value,
                'jgi_proposal_id' : emsl_jgi_dict.get(emsl_proposal[x].value)
                
        }

        #nt = namedtuple('data', data.keys())(*data.values())
        
        #print(nt.data_id, nt.emsl_proposal_id, nt.jgi_proposal_id, dataset_name[x].value) 

        data_dict[dataset_name[x].value] = data

    return data_dict

def get_mapping(filepath):
    
    wb = load_workbook(filename = filepath)
    emsl_jgi_dict = get_emsl_jgi_mapping(wb)
    dataset_mapping = get_data_mapping(wb, emsl_jgi_dict)

    return dataset_mapping