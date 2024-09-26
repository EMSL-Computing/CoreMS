

import re
from typing import Dict, List, Tuple
import sys

sys.path.append("./")
import warnings
warnings.filterwarnings("ignore")

import numpy as np
from pandas.core.frame import DataFrame

from pathlib import Path
import glob

import matplotlib.pyplot as plt
import pandas as pd 
# from PySide2.QtWidgets import QFileDialog, QApplication
# from PySide2.QtCore import Qt

from openpyxl import load_workbook
import logging

from corems.molecular_id.search.molecularFormulaSearch import SearchMolecularFormulas, SearchMolecularFormulasLC
from corems.encapsulation.factory.parameters import LCMSParameters, MSParameters
from corems.molecular_formula.factory.MolecularFormulaFactory import MolecularFormula
from corems.mass_spectra.calc.MZSearch import MZSearch
from corems.molecular_formula.input.masslist_ref import ImportMassListRef
from corems.encapsulation.constant import Labels
from corems.mass_spectra.input import rawFileReader
from support_code.filefinder import get_dirname, get_filename

def _write_frame_to_new_sheet(path_to_file=None, sheet_name='sheet', data=None):
    
    book = None
    try:
    
        book = load_workbook(path_to_file)
    
    except Exception:
    
        logging.debug('Creating new workbook at %s', path_to_file)
    
    with pd.ExcelWriter(path_to_file, engine='openpyxl') as writer:
    
        if book is not None:
            writer.book = book
        
        data_frame = DataFrame(data)

        if not book:
            
            data_frame.to_excel(writer, sheet_name, index=False)
        
        else:
            
            writer.sheets = {ws.title: ws for ws in book.worksheets}

            if sheet_name in writer.sheets:
                
                startrow= writer.sheets[sheet_name].max_row
            
            else:
                startrow=0

            header = False if startrow > 0 else True
            
            data_frame.to_excel(writer, sheet_name, startrow=startrow, index=False, header=header)

def run_thermo(file_location, target_mzs: List[float]) -> Tuple[rawFileReader.DataDependentLCMS, rawFileReader.ImportDataDependentThermoMSFileReader]:
    
    print(file_location)
    
    LCMSParameters.lc_ms.start_scan = -1
    LCMSParameters.lc_ms.end_scan = -1

    LCMSParameters.lc_ms.smooth_window = 5
    
    LCMSParameters.lc_ms.min_peak_datapoints = 3
    LCMSParameters.lc_ms.peak_height_min_percent = 1

    LCMSParameters.lc_ms.eic_signal_threshold = 0.1
    LCMSParameters.lc_ms.eic_tolerance_ppm = 5
    LCMSParameters.lc_ms.enforce_target_ms2 = False
    LCMSParameters.lc_ms.average_target_mz = False
    
    parser = rawFileReader.ImportDataDependentThermoMSFileReader(file_location, target_mzs)
    
    lcms_obj = parser.get_lcms_obj()
                                      
    #ms2_tic, ax_ms2_tic = parser.get_tic(ms_type='MS2', peak_detection=False, plot=False)

    #print(data)

    # get selected data dependent mzs 
    
    
    #plt.legend()
    #plt.tight_layout()
    #plt.title("")
    #plt.show()
    #ax_eic.plot(tic_data.time, tic_data.tic, c='black')
    return lcms_obj, parser
    

def read_lib(ref_filepath:Path):

    ion_charge = -1   

    iontypes = [Labels.protonated_de_ion]
    
    mf_references_dict = ImportMassListRef(ref_filepath).from_lcms_lib_file(ion_charge, iontypes)

    return mf_references_dict

def single_process(mf_references_dict: Dict[str, Dict[float, List[MolecularFormula]]], datapath: Path, current_mix: str, mf_results_dic: dict):

    plt.rcParams["figure.figsize"] = (16,8)

    #get target compounds mz and molecular formulas
    dict_tarrget_mzs = mf_references_dict.get(current_mix)   
    
    target_mzs = dict_tarrget_mzs.keys()

    lcms_obj, parser = run_thermo(datapath, target_mzs)

    target_mzs = parser.selected_mzs
    
    #TODO need to convert this to a lcms object
    scan_number_mass_spectrum = {}
    
    results_list = []
    # mz is from calculate mz
    
    tic_data, ax_tic = lcms_obj.get_tic(ms_type='MS !d', peak_detection=True, 
                                      smooth=True, plot=False)
                                      
    
    eics_data, ax_eic = lcms_obj.get_eics(tic_data,
                                        smooth=True,
                                        plot=False,
                                        legend=False,
                                        peak_detection=True,
                                        ax=ax_tic)
                                        
    lcms_obj.process_ms1(dict_tarrget_mzs)

    #_write_frame_to_new_sheet(path_to_file="HILIC NEG Results.xlsx", sheet_name='all_eic_results', data=results_list)
    # TODO: create lcms and add dependent scans based on scan number 
    # Add Adducts search, right now only working for de or protonated species
    # Export function with csv files
    
    precision_decimals = 0

    ms_peaks_assigned = SearchMolecularFormulasLC(lcms_obj).run_target_worker_ms1()
    
    for eic_peak in lcms_obj:
        
        dependent_scans = parser.iRawDataPlus.GetScanDependents(eic_peak.apex_scan, precision_decimals)

        mass_spectcrum_obj = eic_peak.mass_spectrum
        
        percursordata = {}

        for scan_dependent_detail in dependent_scans.ScanDependentDetailArray:
            
            for precursor_mz in scan_dependent_detail.PrecursorMassArray:
                
                percursordata[precursor_mz] = scan_dependent_detail.ScanIndex
        
        #print(scan, [(mf.name, mf.mz_calc) for mf in mf_references_list], percursordata)
        #print()
        #print(scan, mass_spectcrum_obj.retention_time)
        #print(mf_references_list)
        #SearchMolecularFormulas(mass_spectcrum_obj).run_worker_ms1()
        
        #for precursor_mz in percursordata.keys():
        
        #ax = mass_spectcrum_obj.plot_mz_domain_profile() 
        is_assigned = False
        #target_title = 'Target Molecule(s) = '
        
        #for peak in mass_spectcrum_obj:
            
        #    for mf in peak:
        #        is_assigned = True
                
        #        if not mf.is_isotopologue:
        #            target_title += "{}-{} m/z = {:.4f}".format(mf.name, mf.string_formated, mf.protonated_mz)
                
        #        annotation = "Mol. Form = {}\nm\z = {:.4f}\nerror = {:.4f}\nconfidence score = {:.2f}\nisotopologue score = {:.2f}".format(mf.string_formated, peak.mz_exp, mf.mz_error, mf.confidence_score, mf.isotopologue_similarity)
                
        #        ax.annotate(annotation , xy=(peak.mz_exp, peak.abundance),
        #                                    xytext=(+3, np.sign(peak.abundance)*-40), textcoords="offset points",
        #                                    horizontalalignment="left",
        #                                    verticalalignment="bottom" if peak.abundance > 0 else "top")

        
        #if is_assigned:
            
        #    dir = Path(str(datapath.parent).replace('RAW Files', 'Results MS2 Noise Threshould'))
        #    if not dir.exists():
        #        dir.mkdir(parents=True, exist_ok=True)

        #    ms1_output_file = '{}_{}_{}'.format(scan, 'MS1', datapath.stem)

        #    ax.set_title("Retention Time = {:.3f} {}".format(mass_spectcrum_obj.retention_time, target_title), fontsize=9,)
        #    plt.tight_layout()
        #    #plt.show()
        #    plt.savefig(str(dir) + '/' + ms1_output_file + '.png')
        #    plt.clf()

           
        
        #    mass_spectcrum_obj.to_csv(str(dir) + '/' + ms1_output_file) 
        
        #else:
            
        #    plt.clf()

        scan = eic_peak.apex_scan
        for peak in mass_spectcrum_obj:
            
            for mf in peak:
                
                if not mf.is_isotopologue:
                
                        #error = MZSearch.calc_mz_error(mf.mz_calc, precursor_mz)

                        #check_error = MZSearch.check_ppm_error(LCMSParameters.lcms_obj.eic_tolerance_ppm, error)
                        
                        #if check_error:
                        print(scan, mass_spectcrum_obj.retention_time, mf.name, mf.mz_calc, mf.mz_error, mf.confidence_score, mf.isotopologue_similarity)  
                        #print(peak.mz_exp, precursor_mz, percursordata.get(peak.mz_exp))
                        
                        dependent_scans = parser.iRawDataPlus.GetScanDependents(scan, precision_decimals)
                        
                        selected_for_ms2 = False
                        
                        for scan_dependent_detail in dependent_scans.ScanDependentDetailArray:
                            
                            for index, precursor_mz in enumerate(scan_dependent_detail.PrecursorMassArray):
                                
                                error_ppm_window = (scan_dependent_detail.IsolationWidthArray[index]/precursor_mz) *1000000

                                error = MZSearch.calc_mz_error(mf.mz_calc, precursor_mz)

                                check_error = MZSearch.check_ppm_error(error_ppm_window, error)

                                if  check_error:
                                    
                                    selected_for_ms2 = True
                                    
                                    print(precursor_mz,scan_dependent_detail.ScanIndex, scan_dependent_detail.IsolationWidthArray[index],  scan_dependent_detail.FilterString)
                            
                                    parser.chromatogram_settings.start_scan = scan_dependent_detail.ScanIndex
                                    parser.chromatogram_settings.end_scan = scan_dependent_detail.ScanIndex
                                    
                                    ms2_mass_spec = parser.get_centroid_msms_data(scan_dependent_detail.ScanIndex)
                                    ax = ms2_mass_spec.plot_mz_domain_profile()
                                    
                                    ax.set_title("Retention Time = {:.2f}, Precursor m/z = {:.4f}, Isolation window m/z = {:.1f} \
                                                 Target Molecule = {} m/z = {:.4f} Molecular formula {}\n  ".format(eic_peak.retention_time,
                                                                                                                precursor_mz, scan_dependent_detail.IsolationWidthArray[index],
                                                                                                                mf.name, mf.mz_calc, mf.string_formated), fontsize=9,)
                                                                                                    
                                    #ms_peaks_assigned = SearchMolecularFormulas(mass_spectcrum_obj).search_mol_formulas( mf_references_list, ion_type, find_isotopologues=True)
                                    used_atoms = {'C' : (1, mf.get('C')), 'H': (1, mf.get('H')) }    
                                    
                                    for atoms, value in mf.class_dict.items():
                                        used_atoms[atoms] = (0, value)
                                    
                                    print(used_atoms)
                                    
                                    ms2_mass_spec.molecular_search_settings.usedAtoms = used_atoms
                                    ms2_mass_spec.molecular_search_settings.min_ppm_error = -15 #parser.chromatogram_settings.eic_tolerance_ppm
                                    ms2_mass_spec.molecular_search_settings.max_ppm_error = 15 #parser.chromatogram_settings.eic_tolerance_ppm
                                    ms2_mass_spec.molecular_search_settings.use_min_peaks_filter = False
                                    ms2_mass_spec.molecular_search_settings.use_runtime_kendrick_filter = False
                                    ms2_mass_spec.molecular_search_settings.min_hc_filter = -np.inf
                                    ms2_mass_spec.molecular_search_settings.max_hc_filter = np.inf

                                    ms2_mass_spec.molecular_search_settings.min_oc_filter = -np.inf
                                    ms2_mass_spec.molecular_search_settings.max_oc_filter = np.inf
                                    
                                    ms2_mass_spec.molecular_search_settings.isRadical = False
                                    SearchMolecularFormulas(ms2_mass_spec, find_isotopologues=False).run_worker_mass_spectrum()

                                    fragment_mz = []
                                    fragment_formulas = []
                                    fragment_error = []
                                    cumulative_neutral_loss = []

                                    for msmspeak in ms2_mass_spec:
                                        
                                        for mf_msms in msmspeak:
                                            
                                            fragment_mz.append(round(msmspeak.mz_exp,6))
                                            fragment_formulas.append(mf_msms.string)
                                            fragment_error.append(mf_msms.mz_error)
                                            cumulative_neutral_loss.append(mf.subtract_formula(mf_msms))

                                            annotation = "{} {:.4f}".format(mf_msms.string, mf_msms.mz_error)
                                            ax.annotate(annotation , xy=(msmspeak.mz_exp, msmspeak.abundance),
                                                xytext=(-3, np.sign(msmspeak.abundance)*-3), textcoords="offset points",
                                                horizontalalignment="left",
                                                verticalalignment="bottom" if msmspeak.abundance > 0 else "top")
                                            print(mf_msms, mf_msms.mz_error, mf.subtract_formula(mf_msms))

                                    ms2_output_file = '{}_{}_{}'.format(scan_dependent_detail.ScanIndex, 'MS2', datapath.stem)

                                    result = {'Mix Name': current_mix, 'Data Set': datapath.stem, 'Compound Name': mf.name, 
                                        'MS1 Scan': scan, 'Retention Time': mass_spectcrum_obj.retention_time, 
                                        'm/z': peak.mz_exp, 'm/z Calculated': mf.mz_calc, 'Mol. Formula' : mf.string,  'm/z Error': mf.mz_error, 'Ion Type': mf.ion_type, 
                                        'Confidence Score':  mf.confidence_score, 'Isotopologue Score': mf.isotopologue_similarity, 'm/z Precursor': precursor_mz, 
                                        'Isolation Window': scan_dependent_detail.IsolationWidthArray[index], 'MS2 Scan': scan_dependent_detail.ScanIndex, 
                                        'MS2 m/z': fragment_mz, 'MS2 Mol. Formulas': fragment_formulas, 'MS2 m/z error':fragment_error, 'Cumulative Neutral Loss': cumulative_neutral_loss,
                                        'MS1 Output': 'ms1_output_file', 'MS2 Output': ms2_output_file}
                                    
                                    
                                    dir = Path(str(datapath.parent).replace('RAW Files', 'Results MS2 Noise Threshould'))
                                    
                                    if not dir.exists():
                                        dir.mkdir(parents=True, exist_ok=True)
                                    
                                    ms2_mass_spec.to_csv(str(dir) + '/' + ms2_output_file) 
                
                                    if mf.name not in mf_results_dic.keys():

                                        mf_results_dic[mf.name] = [result]
                                    
                                    else:    
                                        
                                        mf_results_dic[mf.name].append(result)

                                            
                                    plt.tight_layout()
                                    plt.savefig(str(dir) + '/' + ms2_output_file+'.png')
                                    #plt.show() 
                                    plt.clf()
                                     

                        # save results without the fragmentation
                        if not selected_for_ms2:
                            
                            result = {'Mix Name': current_mix, 'Data Set': datapath.stem, 'Compound Name': mf.name, 
                                              'MS1 Scan': scan, 'Retention Time': mass_spectcrum_obj.retention_time, 
                                              'm/z': peak.mz_exp, 'm/z Calculated': mf.mz_calc, 'Mol. Formula' : mf.string,  'm/z Error': mf.mz_error, 'Ion Type': mf.ion_type, 
                                              'Confidence Score':  mf.confidence_score, 'Isotopologue Score': mf.isotopologue_similarity, 'm/z Precursor': None, 
                                              'Isolation Window': None, 'MS2 Scan': None,
                                              'MS2 m/z': None, 'MS2 Mol. Formulas': None, 'MS2 m/z error':None, 'Cumulative Neutral Loss': None, 
                                              'MS1 Output': 'ms1_output_file', 'MS2 Output': None}
                                                
                        
                            if mf.name not in mf_results_dic.keys():

                                mf_results_dic[mf.name] = [result]
                            
                            else:    
                                
                                mf_results_dic[mf.name].append(result)

    return mf_results_dic

    for molecule_name, data in dict_res.items():

        _write_frame_to_new_sheet(path_to_file= 'C18 Results.xlsx', sheet_name='molecular_formula_results', data=data)

        #if  ms_peaks_assigned:
            
            #print(mass_spectcrum_obj.retention_time)
        #    for peak in ms_peaks_assigned:
        #        for mf in peak:
        #            print(peak.mz_exp, mf.mz_calc)

            #mass_spectcrum_obj.to_csv("{:.2f}".format(mass_spectcrum_obj.retention_time).replace('.', "-"))
        #    print()
        #    print([(i.ScanIndex, [i for i in i.PrecursorMassArray]) for i in dependent_scans.ScanDependentDetailArray])
        
        #print()       
        #print(ms_peaks_assigned)
        

def auto_process(mf_references_dict: Dict[str, Dict[float, List[MolecularFormula]]], datadir: Path):

    rootdir = datadir
    
    dict_res = {}

    for mix_name in mf_references_dict.keys():
        
        file_locations = glob.glob(str(rootdir) + "/*.raw")
        file_paths = []
        
        for file_path in file_locations:
            
            mixname_file = re.findall(r'Mix[0-9]{1,2}', file_path)
            if not mixname_file:
                continue
            if mix_name == mixname_file[0]:
                print(mix_name)
                file_paths.append(Path(file_path))
        
        
        if file_paths:
            for file_path in file_paths:
                
                print('***************************' + str(file_path) + '************************')
                try:
                    dict_res = single_process(mf_references_dict, file_path, mix_name, dict_res)
                except:
                    print("PROBLEM FILE {}".format(datadir))
                
        else:
            logging.warn("Could not find a any raw data with mix name: {}".format(mix_name))

    for molecule_name, data in dict_res.items():

        _write_frame_to_new_sheet(path_to_file= 'HILIC NEG Results.xlsx', sheet_name='molecular_formula_results', data=data)

if __name__ == "__main__":
    
    dirpath = "/Users/eber373/OneDrive - PNNL/Documents/Data/LCMS/RAW Files/C18/1st run/NEG/"
    ref_dirpath = "/Users/eber373/OneDrive - PNNL/Documents/Data/LCMS/"
    filename = "LCMS_5191_CapDev_C18_Mix1_NEG_28Apr2021.raw"
    ref_filename = "LCMS_StantardLibrary.csv"
    # run_multiprocess()
    # cpu_percents = monitor(target=run_multiprocess)
    # print(cpu_percents)
    # file_location = get_filename()
    file_location = Path(dirpath + filename)
    ref_file_location = Path(ref_dirpath + ref_filename)
    
    #get the list of molecular formulas objects for each stardard compound maped by mix name
    
    mf_references_dict = read_lib(ref_file_location)
    #loop trought a directory, find and match mix name to raw file, 
    # search eic, do peak picking, for target compounds, currently enforcing parent ion selected to MS2, 
    # to change settings, chage  LCMSParameters.lc_ms parameters inside run_thermo() function:
        
    #auto_process(mf_references_dict, dirpath)
    #mf_results_dic = {}
    if file_location:
        
        #get mix name from filename
        current_mix = (re.findall(r'Mix[0-9]{1,2}', str(file_location)))[0]

        
        single_process(mf_references_dict, file_location, current_mix, mf_references_dict)
        #loop trought a directory, find and match mix name to raw file, 
        # search eic, do peak picking, for target compounds, currently enforcing parent ion selected to MS2, 
        # to change settings, chage  LCMSParameters.lc_ms parameters inside run_thermo() function:
        # 
        #run_thermo(file_location)
