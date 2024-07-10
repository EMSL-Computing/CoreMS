from dataclasses import dataclass
import json
import os
from pathlib import Path
import shutil
import argparse
from zipfile import ZipFile

from openpyxl import load_workbook

from nom_workflow import run_nmdc_workflow
import nmdc_metadata_gen

@dataclass
class EMSL_Metadata:
    data_path: str
    dms_dataset_id: str
    myemsl_link: str
    sample_name: str
    sample_type: str
    env_medium: str
    habitat: str
    ecosystem_category: str
    name: str
    geo_loc_name: str
    lat_long: str
    latitude: float
    longitude: float
    location: str
    ecosystem_type: str
    ecosystem: str
    env_broad_scale: str
    env_local_scale: str
    sample_collection_site: str
    samp_name: str
    ecosystem_subtype: str
    description: str
    collection_date: str
    nmdc_study: str
    biosample_id: str
    
def parse_metadata(metadata_file_path:Path) -> EMSL_Metadata:
        
        wb = load_workbook(filename=metadata_file_path)

        first_sheet = wb.sheetnames[0]

        full_list_worksheet = wb[first_sheet]

        data_name = full_list_worksheet['A']
        dms_dataset_id = full_list_worksheet['B']
        myemsl_link  = full_list_worksheet['C']
        sample_name = full_list_worksheet['D']
        sample_type = full_list_worksheet['E']
        env_medium  = full_list_worksheet['F']
        habitat = full_list_worksheet['G']
        ecosystem_category  = full_list_worksheet['H']
        name  = full_list_worksheet['J']
        geo_loc_name  = full_list_worksheet['K']
        lat_long =  full_list_worksheet['L']
        latitude  = full_list_worksheet['M']
        longitude  = full_list_worksheet['N']
        env_local_scale = full_list_worksheet['O']
        location = full_list_worksheet['P']
        ecosystem_type = full_list_worksheet['Q']
        ecosystem = full_list_worksheet['S']
        env_broad_scale = full_list_worksheet['T']
        sample_collection_site = full_list_worksheet['V']
        samp_name = full_list_worksheet['W']
        ecosystem_subtype = full_list_worksheet['X']
        description = full_list_worksheet['Z']
        collection_date = full_list_worksheet['AA']
        nmdc_study = full_list_worksheet['AB']
        biosample_id = full_list_worksheet['AC']
                        
        for x in range(1, len(full_list_worksheet['A'])):

            metadata = EMSL_Metadata(data_path = Path(data_name[x].value), 
                                dms_dataset_id = dms_dataset_id[x].value,
                                myemsl_link = myemsl_link[x].value,
                                sample_name = sample_name[x].value,
                                sample_type = sample_type[x].value,
                                env_medium  = env_medium[x].value,
                                habitat = habitat[x].value,
                                ecosystem_category  = ecosystem_category[x].value,  
                                name  = name[x].value,
                                geo_loc_name  = geo_loc_name[x].value,
                                lat_long =  lat_long[x].value,
                                latitude  = latitude[x].value,
                                longitude  = longitude[x].value,
                                env_local_scale = env_local_scale[x].value,
                                location = location[x].value,
                                ecosystem_type = ecosystem_type[x].value,
                                ecosystem = ecosystem[x].value,
                                env_broad_scale = env_broad_scale[x].value,
                                sample_collection_site = sample_collection_site[x].value,
                                samp_name = samp_name[x].value,
                                ecosystem_subtype = ecosystem_subtype[x].value,
                                description = description[x].value,
                                collection_date = collection_date[x].value,
                                nmdc_study = nmdc_study[x].value,
                                biosample_id = biosample_id[x].value
                                )
            
            yield metadata

def run_nom_nmdc_data_processing():

    # set command line arguments
    parser = argparse.ArgumentParser(description="A program to run the nmdc nom workflow and create the corresponding metadata objects.")
    parser.add_argument('--data_dir', 
                        '-d', 
                        type=str, 
                        help="The directory path where the raw data lives.")
    parser.add_argument('--metadata_path',
                        '-m',
                        type=str,
                        help="The .xlsx file path where the biosample metadata lives (include .xlsx).")
    parser.add_argument('--registration_file_name',
                        '-rf',
                        type=str,
                        help="The desired name of the registration file (include .json). E.g. emsl_only_grow.json")
    parser.add_argument("--ref_calibration_path",
                        '-r',
                        type=str,
                        help="The .ref path where the reference calibration data lives (include .ref).")
    

    args = parser.parse_args()
    
    file_ext = '.d' 
    data_dir = Path(args.data_dir)
    metadata_file_path = Path(args.metadata_path)

    raw_dir_zip = data_dir / Path("raw_zip/")
    raw_dir_zip.mkdir(parents=True, exist_ok=True)

    results_dir = data_dir / Path("results/")
    results_dir.mkdir(parents=True, exist_ok=True)
    
    failed_files = results_dir / "nom_failed_files.json"
    
    registration_dir = data_dir / 'registration'
    registration_dir.mkdir(parents=True, exist_ok=True)
    registration_file = registration_dir / args.registration_file_name
    
    field_strength = 7
    
    ref_calibration_path = Path(args.ref_calibration_path)
    failed_list = []
    
    nmdc_database = nmdc_metadata_gen.start_nmdc_database()

    for each_data in parse_metadata(metadata_file_path):

        raw_file_path = data_dir / each_data.data_path.with_suffix(file_ext)    

        print(raw_file_path)
        
        issue, ms = run_nmdc_workflow((raw_file_path, ref_calibration_path, field_strength))
        
        try:
            if ms:

                if raw_file_path.suffix =='.d':
                    #raw_file_to_upload_path = zip_d_folder(raw_file_path)    
                    raw_file_to_upload_path = Path(raw_dir_zip / raw_file_path.stem)
                    shutil.make_archive(raw_file_to_upload_path , 'zip', raw_file_path)
                else:
                    raw_file_to_upload_path = raw_file_path

                result_file_name = Path(raw_file_path.name)
                output_file_path = results_dir / result_file_name.with_suffix('.csv')
                ms.to_csv(output_file_path, write_metadata=False)
                
                nmdc_metadata_gen.create_nmdc_metadata(raw_file_to_upload_path.with_suffix('.zip'), 
                                                        output_file_path,
                                                        "https://nmdcdemo.emsl.pnnl.gov/", 
                                                        nmdc_database,
                                                        each_data, biosample_id=each_data.biosample_id)

            else:
                
                print(issue)
                failed_list.append(str(raw_file_path))

        except Exception as inst:

            print(type(inst))    # the exception instance
            print(inst.args)     # arguments stored in .args
            print(inst)
            failed_list.append(str(raw_file_path))
    
    nmdc_metadata_gen.dump_nmdc_database(nmdc_database, registration_file)    

    with failed_files.open('w+') as json_file:

        json_file.write(json.dumps(failed_list, indent=1))

if __name__ == "__main__":

    # run_multiprocess()
    # cpu_percents = monitor(target=run_multiprocess)
    # print(cpu_percents)
    run_nom_nmdc_data_processing()
    # file_location = get_filename()
    # if file_location:
    #    run_assignment(file_location)

