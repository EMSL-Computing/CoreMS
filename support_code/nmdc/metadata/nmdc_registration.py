''' NMDC Schema

 "ChemicalEntity": {
         "additionalProperties": false,
         "description": "An atom or molecule that can be represented with a chemical formula. Include lipids, glycans, natural products, drugs. There may be different terms for distinct acid-base forms, protonation states",
         "properties": {
            "chemical_formula": {
               "description": "A generic grouping for miolecular formulae and empirican formulae",
               "type": "string"
            },
            "has_raw_value": {
               "description": "The value that was specified for an annotation in raw form, i.e. a string. E.g. \"2 cm\" or \"2-4 cm\"",
               "type": "string"
            },
            "inchi": {
               "type": "string"
            },
            "inchi_key": {
               "type": "string"
            },
            "smiles": {
               "description": "A string encoding of a molecular graph, no chiral or isotopic information. There are usually a large number of valid SMILES which represent a given structure. For example, CCO, OCC and C(O)C all specify the structure of ethanol.",
               "items": {
                  "type": "string"
               },
               "type": "array"
            },
            "term": {
               "$ref": "#/definitions/OntologyClass",
               "description": "pointer to an ontology class"
            },
            "was_generated_by": {
               "type": "string"
            }
         },
         "required": [
            "inchi"
         ],
         "title": "ChemicalEntity",
         "type": "object"
      },

"MetaboliteQuantification": {
         "additionalProperties": false,
         "description": "This is used to link a metabolomics analysis workflow to a specific metabolite",
         "properties": {
            "highest_similarity_score": {
               "description": "TODO: Yuri to fill in",
               "type": "number"
            },
            "metabolite_quantified": {
               "description": "the specific metabolite identifier",
               "type": "string"
            }
         },
         "required": [],
         "title": "MetaboliteQuantification",
         "type": "object"
      },

"MetabolomicsAnalysisActivity": {
         "additionalProperties": false,
         "description": "",
         "properties": {
            "activity_id": {
               "type": "string"
            },
            "ended_at_time": {
               "type": "string"
            },
            "execution_resource": {
               "description": "Example: NERSC-Cori",
               "type": "string"
            },
            "git_url": {
               "description": "Example: https://github.com/microbiomedata/mg_annotation/releases/tag/0.1",
               "type": "string"
            },
            "has_input": {
               "description": "An input to a process.",
               "items": {
                  "type": "string"
               },
               "type": "array"
            },
            "has_metabolite_quantifications": {
               "items": {
                  "$ref": "#/definitions/MetaboliteQuantification"
               },
               "type": "array"
            },
            "has_output": {
               "description": "An output biosample to a processing step",
               "items": {
                  "type": "string"
               },
               "type": "array"
            },
            "started_at_time": {
               "type": "string"
            },
            "used": {
               "description": "The instrument used to collect the data used in the analysis",
               "type": "string"
            },
            "was_associated_with": {
               "$ref": "#/definitions/Agent"
            },
            "was_informed_by": {
               "type": "string"
            }
         },
         "required": [
            "activity_id"
         ],
         "title": "MetabolomicsAnalysisActivity",
         "type": "object"
      },
'''

from datetime import datetime
from pathlib import Path
import hashlib
import uuid
import json

from openpyxl import load_workbook

class DMS_Mapping():

    def __init__(self, dms_file_path) -> None:

        self.dms_file_path = Path(dms_file_path)

    @staticmethod
    def get_emsl_jgi_mapping(wb):

        last_sheet = wb.sheetnames[-1]
        emsl_jgi_mapping = wb[last_sheet]
        emsl_proposals = emsl_jgi_mapping['B']
        jgi_proposals = emsl_jgi_mapping['A']
        emsl_jgi_dict = {}
        for x in range(len(emsl_proposals)):
            emsl_jgi_dict[emsl_proposals[x].value] = jgi_proposals[x].value
        return emsl_jgi_dict

    @staticmethod
    def get_data_mapping(wb, emsl_jgi_dict):

        first_sheet = wb.sheetnames[2]

        full_list_worksheet = wb[first_sheet]

        emsl_proposal = full_list_worksheet['AP']
        dataset_id = full_list_worksheet['A']
        dataset_name = full_list_worksheet['B']
        experiment_id = full_list_worksheet['AT']
        instrument_name = full_list_worksheet['E']
        data_dict = {}

        for x in range(len(dataset_id)):

            data = {
                'data_id': dataset_id[x].value,
                'experiment_id': experiment_id[x].value,
                'emsl_proposal_id': emsl_proposal[x].value,
                'instrument_name': instrument_name[x].value
                #'jgi_proposal_id': emsl_jgi_dict.get(emsl_proposal[x].value)
            }

            # nt = namedtuple('data', data.keys())(*data.values())

            # print(nt.data_id, nt.emsl_proposal_id, nt.jgi_proposal_id, dataset_name[x].value)

            data_dict[dataset_name[x].value] = data

        return data_dict

    def get_selected_sample_list(self):

        wb = load_workbook(filename=self.dms_file_path)

        second_sheet = wb.sheetnames[2]

        # print(second_sheet)
        full_list_worksheet = wb[second_sheet]

        sample_name = full_list_worksheet['B']

        instrument_name = full_list_worksheet['E']

        for x in range(1, len(sample_name)):

            yield Path(sample_name[x].value), int(instrument_name[x].value.strip()[0:2])

    def get_mapping(self):

        wb = load_workbook(filename=self.dms_file_path)
        "removing mapping"
        emsl_jgi_dict = {}
        dataset_mapping = self.get_data_mapping(wb, emsl_jgi_dict)

        return dataset_mapping

class NMDC_Metadata:

    def __init__(self, in_file_path, calibration_file_path, out_file_path, dms_file_path) -> None:

        self.dataset_mapping = DMS_Mapping(dms_file_path).get_mapping()
        self.in_file_path = Path(in_file_path)
        self.calibration_file_path = calibration_file_path
        self.out_file_path = Path(out_file_path)
        
    @staticmethod
    def get_mds_metadata(dataset_name):
        ''' data: [{'#Row': '1', 
                    'ID': 906829,
                    'Dataset': 'ABF_Rt_43016_RNAi_48h_14152_R2_47_02_13April21_Romulus_WBEH-21-01-09',
                    'Experiment': 'ABF_Rt_43016_RNAi_48h_14152_R2_47_02',
                    'Campaign': 'Agile',
                    'State': 'Complete',
                    'Instrument': 'QEHFX03',
                    'Created': '2021-04-14 17:46:39.200',
                    'Comment': '',
                    'Rating': 'Released',
                    'Dataset Type': 'HMS-HCD-HMSn',
                    'Operator': 'ATTA556',
                    'Dataset Folder Path': '\\\\proto-8\\QEHFX03\\2021_2\\ABF_Rt_43016_RNAi_48h_14152_R2_47_02_13April21_Romulus_WBEH-21-01-09',
                    'QC_Link': 'https://Proto-8.pnl.gov/QEHFX03/2021_2/ABF_Rt_43016_RNAi_48h_14152_R2_47_02_13April21_Romulus_WBEH-21-01-09/QC/index.html',
                    'Acq Start': '2021-04-14 15:42:45.000',
                    'Acq. End': '2021-04-14 17:42:45.000',
                    'Acq Length': 120,
                    'Scan Count': 61752,
                    'File Size MB': '1654.73',
                    'Cart Config': 'Romulus_WatersAQ_Standard_Jan2019',
                    'LC Column': 'WBEH-21-01-09',
                    'Separation Type': 'LC-Dionex-Formic_2hr',
                    'Request': 959190, 
                    'Work Package': 'XXXX',
                    'Organism': 'Rhodosporidium_toruloides',
                    'Tissue': None,
                    '#DateSortKey': '2021-04-14 15:42:45.000'}] 
        '''
        username = os.environ.get('USER_NAME') or 'pnnl'
        password = os.environ.get('USER_PASSWORD') or 'pnnl_password'

        url_request = "https://dms2.pnl.gov/data/ax/json/list_report/dataset/{}".format(dataset_name) 
        r = requests.get(url_request, auth=(username, password))
        data = r.json()

        return data

    def save_nmdc_metadata(self,
                           data_obj,
                           nom=False,
                           registration_path="gcms_metabolomics_data_products.json",
                           bucket="metabolomics",
                           repo_url="https://github.com/microbiomedata/metaMS"):

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data_id = self.get_dataid(self.in_file_path, self.dataset_mapping)
        output_id = self.get_md5(self.out_file_path)
        was_informed_by = self.get_was_informed_by(self.in_file_path, self.dataset_mapping)
        activity_id = "{}:{}".format("nmdc", uuid.uuid4().hex)
        has_calibration = self.get_dataid(self.calibration_file_path, self.dataset_mapping)

        metabolomics_analysis_activity = {

            "id": activity_id,
            "type": "ndmc:NomAnalysisActivity",
            "ended_at_time": now,
            "execution_resource": "EMSL-RZR",
            "git_url": repo_url,
            "has_input": [data_id],
            "has_calibration": has_calibration,
            "has_output": [output_id],
            "started_at_time": now,
            "used": self.get_instrument_name(self.in_file_path, self.dataset_mapping),
            "was_informed_by": was_informed_by,
        }

        if not nom:

            metabolomics_analysis_activity["type"] = "ndmc:MetabolomicsAnalysisActivity"
            if data_obj:
                metabolomics_analysis_activity["has_metabolite_quantifications"] = self.get_metabolites_objs(data_obj)
            # metabolomics_analysis_activity["has_mformula_quantifications"] = self.get_metabolites_objs(data_obj)

        with open(self.out_file_path.with_suffix('.json'), 'w') as metadata_output:
            metadata_output.write(json.dumps(metabolomics_analysis_activity, indent=1))

        if not nom:
            object_type = "GC-MS Metabolomics Results"
            description = "MetaMS GC-MS metabolomics output detail CSV file"

            self.add_metabolomics_data_product(output_id, activity_id, bucket=bucket, 
                                           object_type=object_type, registration_path=registration_path,
                                           description=description)

        else:

            self.add_metabolomics_data_product(output_id, activity_id, bucket=bucket, 
                                                registration_path=registration_path)
                                           

    def add_metabolomics_data_product(self, output_id, activity_id, bucket, 
                                      registration_path="gcms_metabolomics_data_products.json",
                                      object_type = "FT ICR-MS Analysis Results",
                                      description="EnviroMS FT ICR-MS natural organic matter workflow molecular formula assignment output details"):

        data_obj = [{
                    "id": output_id,
                    "name": self.out_file_path.name,
                    "description": description,
                    "file_size_bytes": self.out_file_path.stat().st_size,
                    "md5_checksum": hashlib.md5(self.out_file_path.open('rb').read()).hexdigest(),
                    "url": "{}/{}/{}/{}".format("https://nmdcdemo.emsl.pnnl.gov", bucket, "results", self.out_file_path.name),
                    "was_generated_by": activity_id,
                    "type": "nmdc:DataObject",
                    "data_object_type": object_type
                    }]
        
        metabolomics_data_path = Path(registration_path)

        if metabolomics_data_path.exists():

            with metabolomics_data_path.open('r') as metabolomics_data_products:
                products = json.load(metabolomics_data_products)
                products.extend(data_obj)

        else:
            products = data_obj

        with metabolomics_data_path.open('w') as metabolomics_data_products:
            metabolomics_data_products.write(json.dumps(products, indent=1))  

        metadata_url = ["{}/{}/{}/{}".format("https://nmdcdemo.emsl.pnnl.gov", bucket, "metadata", self.out_file_path.with_suffix('.json').name)]
        
        self.dump_metadata_registation(registration_path, metadata_url)

    def dump_metadata_registation(self, registration_path, metadata_url):

        metabolomics_metadata_path = Path(str(registration_path).replace('data', 'metadata')) 
        
        if metabolomics_metadata_path.exists():

            with metabolomics_metadata_path.open('r') as metabolomics_data_products:
                products = json.load(metabolomics_data_products)
                products.extend(metadata_url)

        else:
            products = metadata_url

        with metabolomics_metadata_path.open('w') as metabolomics_data_products:
            metabolomics_data_products.write(json.dumps(products, indent=1))  

    @staticmethod
    def get_instrument_name(in_file_path, dataset_mapping):

        if not in_file_path:
            return None

        mapping = dataset_mapping.get(in_file_path.stem)
        if mapping:
            data_id = mapping.get("instrument_name")
            return "{}".format(data_id)
        else:
            return None

    @staticmethod
    def get_dataid(in_file_path, dataset_mapping):

        if not in_file_path:
            return None

        mapping = dataset_mapping.get(in_file_path.stem)

        if mapping:
            data_id = mapping.get("data_id")
            return "{}:{}_{}".format("emsl", "output", data_id)
        else:
            return None

    @staticmethod
    def get_was_informed_by(in_file_path, dataset_mapping):

        if not in_file_path:
            return None

        mapping = dataset_mapping.get(in_file_path.stem)
        if mapping:
            data_id = mapping.get("data_id")
            return "{}:{}".format("emsl", data_id)
        else:
            return None

    @staticmethod
    def get_md5(out_file_path):
        bytes_io = out_file_path.open('rb').read()

        return "{}:{}".format("nmdc", hashlib.md5(bytes_io).hexdigest())

    @staticmethod
    def get_metabolites_objs(gcms_obj):

        all_metabolites = []

        for metabolite in gcms_obj.metabolites_data:

            metabolite_quantification = {}

            metabolite_quantification["type"] = "ndmc:MetaboliteQuantification"
            metabolite_quantification["highest_similarity_score"] = metabolite.get("highest_similarity_score")
            metabolite_quantification["metabolite_quantified"] = "{}:{}".format("chebi", metabolite.get("chebi"))
            metabolite_quantification["alternate_identifiers"] = ["{}:{}".format("kegg", metabolite.get("kegg")),
                                                                  "{}:{}".format("cas", metabolite.get("casno"))]

            all_metabolites.append(metabolite_quantification)

        return all_metabolites

    def create_nmdc_gcms_metadata(self, gcms_obj, registration_path="gcms_metabolomics_data_products.json"):

        # print(dataset_mapping.get('GCMS_Blank_08_GC-01_20150622'))
        bucket = "metabolomics"
        repo_url = "https://github.com/microbiomedata/metaMS"
        self.save_nmdc_metadata(gcms_obj, nom=False, registration_path=registration_path, bucket=bucket, repo_url=repo_url)

    def create_nmdc_ftms_metadata(self, ms_obj, registration_path="spruce_ftms_nom_data_products.json"):
        # print(dataset_mapping.get('GCMS_Blank_08_GC-01_20150622'))
        bucket = "nom"
        repo_url = "https://github.com/microbiomedata/enviroMS"
        self.save_nmdc_metadata(ms_obj, nom=True, registration_path=registration_path, bucket=bucket, repo_url=repo_url)
